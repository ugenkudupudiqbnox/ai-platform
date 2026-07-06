"""Excel reader component (constitution §15/§16).

Generic, reusable reader for ``.xlsx``/``.xls`` files. Reads the requested
sheet/range with ``openpyxl`` and emits the rows in the §14 envelope
(``data.rows`` = list of dict rows when ``has_header``, else list of lists),
shaped toward ``DocumentManifest``. It is a custom reader rather than the
LangFlow built-in file component so the output conforms to the project's
contracts and the §16 PII rules (see ADR-0002).

``openpyxl`` is baked into ``docker/langflow/Dockerfile`` (see ADR-0004); if it
is not importable the reader returns ``code=AR_NOT_IMPLEMENTED`` rather than
raising. Never raises (§5/§9): file-not-found / corrupt-file / parse errors
surface as ``code=AR_VALIDATION`` envelopes.
"""

import json
import os

from lfx.custom import Component
from lfx.io import BoolInput, IntInput, MessageTextInput, Output
from lfx.schema import Message


def _envelope(status: str, code: str, data: dict | None = None,
              error: dict | None = None) -> dict:
    env: dict = {"status": status, "code": code, "data": data or {}}
    if error:
        env["error"] = error
    return env


class ExcelReaderComponent(Component):
    name = "ExcelReaderComponent"
    display_name = "Excel Reader"
    description = (
        "Read an Excel (.xlsx/.xls) file and return its rows in the canonical "
        "envelope. Call this when the user provides a spreadsheet of invoices, "
        "receipts, or AR data that must be ingested and matched."
    )
    icon = "Sheet"

    inputs = [
        MessageTextInput(
            name="file_path",
            display_name="File Path",
            info="Path to the Excel file (resolved inside the LangFlow container; must be on an SSRF-allowed host if fetched remotely, §16).",
            required=True,
            tool_mode=True,
        ),
        MessageTextInput(
            name="sheet_name",
            display_name="Sheet",
            info="Sheet name to read (defaults to the first/active sheet).",
            tool_mode=True,
        ),
        MessageTextInput(
            name="range",
            display_name="Range",
            info="Optional A1-style range, e.g. 'A1:D200'.",
            tool_mode=True,
        ),
        BoolInput(
            name="has_header",
            display_name="First Row Is Header",
            value=True,
            info="Treat the first row as column headers.",
        ),
        IntInput(
            name="max_rows",
            display_name="Max Rows",
            value=0,
            info="Maximum rows to return (0 = no limit).",
        ),
    ]

    outputs = [
        Output(
            name="reader_output",
            display_name="Rows",
            method="read",
        ),
    ]

    def read(self) -> Message:
        file_path = (self.file_path or "").strip()
        if not file_path:
            return Message(text=json.dumps(_envelope(
                "error", "AR_VALIDATION",
                error={"message": "file_path is required"})))
        if not os.path.isfile(file_path):
            return Message(text=json.dumps(_envelope(
                "error", "AR_VALIDATION",
                error={"message": f"file not found: {file_path}"})))
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            return Message(text=json.dumps(_envelope(
                "error", "AR_NOT_IMPLEMENTED",
                error={"message": f"openpyxl not available in image: {exc}"}
            )))
        sheet_name = (self.sheet_name or "").strip() or None
        rng = (self.range or "").strip() or None
        has_header = bool(self.has_header) if self.has_header is not None else True
        try:
            max_rows = int(self.max_rows or 0)
        except (TypeError, ValueError):
            max_rows = 0
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name else wb.active
            if rng:
                # ws[rng] returns a tuple of tuples of Cell objects.
                cell_rows = ws[rng]
                all_rows = [[("" if c.value is None else str(c.value)) for c in row]
                            for row in cell_rows]
            else:
                all_rows = [[("" if v is None else str(v)) for v in row]
                            for row in ws.iter_rows(values_only=True)]
            wb.close()
        except (KeyError, OSError, ValueError) as exc:
            return Message(text=json.dumps(_envelope(
                "error", "AR_VALIDATION",
                error={"message": f"Excel read failed: {exc}"})))
        except Exception as exc:  # noqa: BLE001 — openpyxl raises generic errors on corrupt files
            return Message(text=json.dumps(_envelope(
                "error", "AR_VALIDATION",
                error={"message": f"Excel open failed: {exc}"})))
        if has_header and all_rows:
            headers = [str(h) for h in all_rows[0]]
            body = all_rows[1:]
            if max_rows:
                body = body[:max_rows]
            rows = [dict(zip(headers, row)) for row in body]
        else:
            rows = all_rows[:max_rows] if max_rows else all_rows
        return Message(text=json.dumps(_envelope(
            "ok", "AR_OK",
            data={"file": file_path, "rows": rows, "row_count": len(rows),
                  "has_header": has_header,
                  "sheet": sheet_name or "active"})))